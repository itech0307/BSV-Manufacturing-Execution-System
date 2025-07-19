import openpyxl
from copy import copy
import pandas as pd
from .models import SalesOrder, ProductionPlan
from datetime import datetime
from django.http import HttpResponse, JsonResponse
from django.conf import settings
from openpyxl.drawing.image import Image
import qrcode
import io
import os

# Celery
from celery import shared_task
# Celery-progress
from celery_progress.backend import ProgressRecorder

import boto3

s3 = boto3.client('s3', region_name=settings.AWS_S3_REGION_NAME)

def get_s3_file(key):
    """Get the file from S3 and return it as a BytesIO object."""
    s3_response = s3.get_object(Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=key)
    file_content = s3_response['Body'].read()
    return io.BytesIO(file_content)

def index2(request):
    pass

def copy_sheet_attributes(source_sheet, target_sheet):
    if isinstance(source_sheet, openpyxl.worksheet._read_only.ReadOnlyWorksheet):
        return
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = copy(source_sheet.freeze_panes)

    # set row dimensions
    # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
    for rn in range(len(source_sheet.row_dimensions)):
        target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])

    if source_sheet.sheet_format.defaultColWidth is None:
        pass
    else:
        target_sheet.sheet_format.defaultColWidth = copy(source_sheet.sheet_format.defaultColWidth)

    # set specific column width and hidden property
    # we cannot copy the entire column_dimensions attribute so we copy selected attributes
    for key, value in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[key].min = copy(source_sheet.column_dimensions[key].min)   # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
        target_sheet.column_dimensions[key].max = copy(source_sheet.column_dimensions[key].max)  # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the issue. Note that this is also the case for the width, not onl;y the hidden property
        target_sheet.column_dimensions[key].width = copy(source_sheet.column_dimensions[key].width) # set width for every column
        target_sheet.column_dimensions[key].hidden = copy(source_sheet.column_dimensions[key].hidden)

def copy_cells(source_sheet, target_sheet):
    for r, row in enumerate(source_sheet.iter_rows()):
        for c, cell in enumerate(row):
            source_cell = cell
            if isinstance(source_cell, openpyxl.cell.read_only.EmptyCell):
                continue
            target_cell = target_sheet.cell(column=c+1, row=r+1)

            target_cell._value = source_cell._value
            target_cell.data_type = source_cell.data_type

            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

            if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)

            if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.comment:
                target_cell.comment = copy(source_cell.comment)

def copy_sheet(source_sheet, target_sheet):
    copy_cells(source_sheet, target_sheet)  # copy all the cel values and styles
    copy_sheet_attributes(source_sheet, target_sheet)

def parse_date(date_string):
    if not date_string or pd.isna(date_string):
        return None
    
    date_formats = ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d']
    for date_format in date_formats:
        try:
            return datetime.strptime(str(date_string), date_format).date()
        except ValueError:
            continue
    
    return None

@shared_task(bind=True)
def ordersheet_upload_celery(self,df_json):
    try:
        # Object for recording progress of the operation
        progress_recorder = ProgressRecorder(self)
        
        # Convert JSON back to pandas DataFrame
        df = pd.read_json(df_json)
        df = df.sort_values(by='Receipt date', ascending=True)
        
        for i in range(len(df)):
            # Skip if Sales order and Line number are empty
            if df['Sales order'][i] == "" and df['Line number'][i] == "":
                continue
        
            sales_order = SalesOrder.objects.filter(
                order_no = f"{df['Sales order'][i]}-{(df['Line number'][i])}"
            ).first()
            
            if sales_order:
                # Compare values in the order_quantity field of the object
                if (int(df['Quantity'][i])) < 0:
                    
                    # Change order_status field to false
                    #sales_order.status = True
                    #sales_order.save()  # Save changes
                    pass
                
                elif (int(df['Quantity'][i])) > 0:
                    # Change the order_status field to None
                    sales_order.status = None
                    sales_order.save()  # Save changes
                    
                    order_data = {
                        'order_id': df['Sales order'][i],
                        'seq_no': int(df['Line number'][i]),
                        'customer_order_no': df['po number'][i],
                        'customer_name': df['Customer Name'][i],
                        'order_type': df['Sales origin'][i],
                        'order_date': parse_date(df['Receipt date'][i]),
                        'rtd': parse_date(df['RTD'][i]),
                        'etd': parse_date(df['ETD'][i]),
                        'brand': df['Brand Name'][i],
                        'item_name': df['Item Name'][i],
                        'color_code': df['Color Code'][i],
                        'color_name': df['Color Name'][i],
                        'pattern': df['TYPE'][i],
                        'spec': df['Spec Name'][i],
                        'order_qty': int(df['Quantity'][i]),
                        'qty_unit': df['Unit'][i],
                        'unit_price': float(df['Ship Unit price'][i]),
                        'currency': df['Currency(Trade)'][i],
                        'order_remark': df['Prod. remark'][i],
                        'model_name': df['Model name'][i],
                        'sample_step': df['Sample Step'][i],
                        'production_location': df['Order To Company'][i],
                        'product_group': df['Prod Group'][i],
                        'product_type': df['Custom No'][i]
                    }
        
                    for key, value in order_data.items():
                        setattr(sales_order, key, value)
                    sales_order.save()
            else:
                if int(df['Quantity'][i]) > 0:
                    order_data = {
                        'order_id': df['Sales order'][i],
                        'seq_no': int(df['Line number'][i]),
                        'customer_order_no': df['po number'][i],
                        'customer_name': df['Customer Name'][i],
                        'order_type': df['Sales origin'][i],
                        'order_date': parse_date(df['Receipt date'][i]),
                        'rtd': parse_date(df['RTD'][i]),
                        'etd': parse_date(df['ETD'][i]),
                        'brand': df['Brand Name'][i],
                        'item_name': df['Item Name'][i],
                        'color_code': df['Color Code'][i],
                        'color_name': df['Color Name'][i],
                        'pattern': df['TYPE'][i],
                        'spec': df['Spec Name'][i],
                        'order_qty': int(df['Quantity'][i]),
                        'qty_unit': df['Unit'][i],
                        'unit_price': float(df['Ship Unit price'][i]),
                        'currency': df['Currency(Trade)'][i],
                        'order_remark': df['Prod. remark'][i],
                        'model_name': df['Model name'][i],
                        'sample_step': df['Sample Step'][i],
                        'production_location': df['Order To Company'][i],
                        'product_group': df['Prod Group'][i],
                        'product_type': df['Custom No'][i]
                    }
                    
                    SalesOrder.objects.create(**order_data)
                else:
                    continue
        
            progress_recorder.set_progress(i + 1, len(df), description="Uploading")
    except Exception as e:
        # Print error message if an error occurs during the operation
        print(f"An error occurred during the operation: {e}")
        # Add necessary exception handling here

def dryplan_convert_to_qrcard(df_json):
    try:
        # Create file name with current time
        current_time = datetime.now().strftime('%y%m%d%H%M%S')
        filename = f"qr_code_{current_time}.xlsx"
        
        # Change Content-Disposition with new file name
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Convert JSON data to pandas DataFrame
        df = pd.read_json(df_json)
        
        # Create new workbook and load existing 'qrcard.xlsx' file
        wb = openpyxl.Workbook()
        
        # Get the excel file from S3
        excel_file = get_s3_file('forms/qrcard.xlsx')

        # Load the workbook using openpyxl
        qrcard = openpyxl.load_workbook(excel_file)
        
        # For each row, create a new worksheet and fill in the data
        for i in range(len(df)):
            ws = wb.create_sheet(str(i+1))
            copy_sheet(qrcard['DRY'], ws)
            # Fill in the data in the worksheet
            row_data = df.iloc[i]  # Preload data from the data frame for one row
            try:
                sales_order = SalesOrder.objects.exclude(status=False).get(order_no=f"{row_data.iloc[0]}-{row_data.iloc[1]}")
                if '/' in str(row_data.iloc[14]):
                    skin_resin, binder_resin = str(row_data.iloc[14]).split('/')
                else:
                    skin_resin, binder_resin = '', ''
                
                pd_information = {
                    "base": f"{row_data.iloc[7]}",
                    "skin_resin": skin_resin,
                    "binder_resin": binder_resin,
                    "rp_qty": f"{row_data.iloc[15]}",
                    "plan_remark": f"{row_data.iloc[16]}"
                }
                
                # Create or update ProductionPlan model instance
                production_plan, created = ProductionPlan.objects.update_or_create(
                    sales_order=sales_order,
                    plan_date=row_data.iloc[11][:10],
                    defaults={
                        'plan_no': f"{row_data.iloc[12]}",
                        'plan_qty': int(row_data.iloc[13]),
                        'pd_line': f"{row_data.iloc[10]}",
                        'item_group': 'Dry',
                        'pd_information': pd_information
                    }
                )
            except:
                pass

            ws['A3'] = row_data.iloc[2]  # customer
            ws['F3'] = row_data.iloc[3]  # brand
            ws['A4'] = row_data.iloc[4]  # item
            ws['A5'] = row_data.iloc[5]  # color
            ws['F5'] = row_data.iloc[6]  # pattern
            ws['A6'] = f"{row_data.iloc[0]}-{row_data.iloc[1]}"   # order number
            ws['G6'] = row_data.iloc[7]  # Base
            ws['D6'] = row_data[8]  # order qty
            ws['C9'] = row_data.iloc[9]  # Remark
            ws['C11'] = row_data.iloc[16]  # Plan Remark

            ws['A17'] = f'D{row_data.iloc[10]}-{row_data.iloc[12]}'  # Line
            ws['F17'] = row_data.iloc[11][5:10]  # plan date
            ws['G7'] = f"Plan Qty: {int(row_data.iloc[13])} M"  # plan qty
            ws['A7'] = row_data.iloc[17]  # order type
            
            qr_str = f"!BSVPD!{row_data.iloc[0]}!{row_data.iloc[1]}!"
            qr_img = qrcode.make(qr_str).resize((160, 160))
            
            # Save QR code image to BytesIO stream
            img_stream = io.BytesIO()
            qr_img.save(img_stream, format='PNG')
            
            # Reset stream position to the beginning
            img_stream.seek(0)
            
            # Create openpyxl image object
            qr_img_openpyxl = Image(img_stream)
            
            # Add QR code image to worksheet
            ws.add_image(qr_img_openpyxl, "G22")

            qr_img_2 = qrcode.make(qr_str).resize((160, 160))

            # Save QR code image to BytesIO stream
            img_stream_2 = io.BytesIO()
            qr_img_2.save(img_stream_2, format='PNG')
            
            # Reset stream position to the beginning
            img_stream_2.seek(0)
            
            # Create openpyxl image object
            qr_img_openpyxl_2 = Image(img_stream_2)
            
            # Add QR code image to worksheet
            ws.add_image(qr_img_openpyxl_2, "A22")
        
        # Save workbook and return response
        del wb['Sheet']
        wb.save(response)
        return response
    
    except Exception as e:
        # Return error message in JSON format if an error occurs
        return JsonResponse({'error': str(e)})

def dev_order_convert_to_qrcard(development_and_orders):
    try:
        # Create excel file response
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="dev_qrcard_.xlsx"'
        
        # Convert JSON data to pandas DataFrame
        # df = pd.read_json(df_json)
        
        # Create new workbook and load existing 'qrcard.xlsx' file
        wb = openpyxl.Workbook()
        
        # Get the excel file from S3
        excel_file = get_s3_file('forms/qrcard.xlsx')

        # Load the workbook using openpyxl
        qrcard = openpyxl.load_workbook(excel_file)
        
        # For each row, create a new worksheet and fill in the data
        for development, order in development_and_orders:
            order_no = order.order_no
            order_info = order.order_information
            ws = wb.create_sheet(str(order_no))
            copy_sheet(qrcard['DEV'], ws)
            # Fill in the data in the worksheet
            
            ws['A3'] = str(development.category)
            #ws['F3'] = str(development.deadline)[2:]
            ws['A4'] = order_info['item']
            ws['A5'] = order_info['color']
            ws['F5'] = order_info['pattern']
            ws['A6'] = str(order_no)
            ws['G6'] = order_info['base']  # Base
            ws['A7'] = order_info['skin_resin']  # Base
            ws['G7'] = order_info['binder_resin']  # Base
            ws['D6'] = order_info['order_qty']  # order qty
            ws['C9'] = str(development.title)  # subject
            ws['C10'] = str(development.purpose)  # purpose
            ws['C11'] = str(development.content)  # Remark

            order_no.split('-')
            
            qr_str = f"!BSVPD!{order_no.split('-')[0]}!{order_no.split('-')[1]}!"
            qr_img = qrcode.make(qr_str).resize((160, 160))
            
            # Save QR code image to BytesIO stream
            img_stream = io.BytesIO()
            qr_img.save(img_stream, format='PNG')
            
            # Reset stream position to the beginning
            img_stream.seek(0)
            
            # Create openpyxl image object
            qr_img_openpyxl = Image(img_stream)
            
            # Add QR code image to worksheet
            ws.add_image(qr_img_openpyxl, "G22")

            qr_img_2 = qrcode.make(qr_str).resize((160, 160))

            # Save QR code image to BytesIO stream
            img_stream_2 = io.BytesIO()
            qr_img_2.save(img_stream_2, format='PNG')
            
            # Reset stream position to the beginning
            img_stream_2.seek(0)
            
            # Create openpyxl image object
            qr_img_openpyxl_2 = Image(img_stream_2)
            
            # Add QR code image to worksheet
            ws.add_image(qr_img_openpyxl_2, "A22")
        
        # Save workbook and return response
        del wb['Sheet']
        wb.save(response)
        return response
    
    except Exception as e:
        # Return error message in JSON format if an error occurs
        return JsonResponse({'error': str(e)})