import openpyxl
from copy import copy
import pandas as pd
from .models import ColorSwatchMovement, ColorSwatch
from production_management.models import SalesOrder, ProductionPlan
from datetime import datetime
from django.http import HttpResponse, JsonResponse
from django.conf import settings
from django.db import transaction
from openpyxl.drawing.image import Image
import qrcode
import io
import os
import logging
import pyxlsb

# Celery
from celery import shared_task
# Celery-progress
from celery_progress.backend import ProgressRecorder
from celery.schedules import crontab
from config.celery import app

# Set up logging
logger = logging.getLogger(__name__)

import boto3

s3 = boto3.client('s3', region_name=settings.AWS_S3_REGION_NAME)

def get_s3_file(key):
    """Get a file from S3 and return it as a BytesIO object.""" 
    s3_response = s3.get_object(Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=key)
    file_content = s3_response['Body'].read()
    return io.BytesIO(file_content)

def index2(request):
    pass

def copy_sheet_attributes(source_sheet, target_sheet):
    if isinstance(source_sheet, openpyxl.worksheet._read_only.ReadOnlyWorksheet):
        return
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = copy(source_sheet.freeze_panes)

    # set row dimensions
    # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
    for rn in range(len(source_sheet.row_dimensions)):
        target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])

    if source_sheet.sheet_format.defaultColWidth is None:
        pass
    else:
        target_sheet.sheet_format.defaultColWidth = copy(source_sheet.sheet_format.defaultColWidth)

    # set specific column width and hidden property
    # we cannot copy the entire column_dimensions attribute so we copy selected attributes
    for key, value in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[key].min = copy(source_sheet.column_dimensions[key].min)   # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
        target_sheet.column_dimensions[key].max = copy(source_sheet.column_dimensions[key].max)  # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the issue. Note that this is also the case for the width, not onl;y the hidden property
        target_sheet.column_dimensions[key].width = copy(source_sheet.column_dimensions[key].width) # set width for every column
        target_sheet.column_dimensions[key].hidden = copy(source_sheet.column_dimensions[key].hidden)

def copy_cells(source_sheet, target_sheet):
    for r, row in enumerate(source_sheet.iter_rows()):
        for c, cell in enumerate(row):
            source_cell = cell
            if isinstance(source_cell, openpyxl.cell.read_only.EmptyCell):
                continue
            target_cell = target_sheet.cell(column=c+1, row=r+1)

            target_cell._value = source_cell._value
            target_cell.data_type = source_cell.data_type

            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

            if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)

            if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.comment:
                target_cell.comment = copy(source_cell.comment)

def copy_sheet(source_sheet, target_sheet):
    copy_cells(source_sheet, target_sheet)  # copy all the cel values and styles
    copy_sheet_attributes(source_sheet, target_sheet)

def parse_date(date_string):
    if not date_string or pd.isna(date_string):
        return None
    
    date_formats = ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d']
    for date_format in date_formats:
        try:
            return datetime.strptime(str(date_string), date_format).date()
        except ValueError:
            continue
    
    return None

def order_convert_to_qrcard(production_order):
    try:
        # Generate a response in an Excel file.
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="qrcard_.xlsx"'
        
        # Convert JSON data to pandas DataFrame
        # df = pd.read_json(df_json)
        
        # Create a new workbook and load the existing 'qrcard.xlsx' file
        wb = openpyxl.Workbook()
        
        # Get the file from S3
        excel_file = get_s3_file('forms/qrcard.xlsx')

        # Load the workbook using openpyxl
        qrcard = openpyxl.load_workbook(excel_file)
        
        # Create a new worksheet for each row and fill the data
        for order in production_order:
            order_no = order.order_no            
            
            ws = wb.create_sheet(str(order_no))
            copy_sheet(qrcard['DRY'], ws)
            # Fill the worksheet with data
            
            ws['A3'] = order.customer_name
            ws['F3'] = order.brand
            ws['A4'] = order.item_name
            ws['A5'] = order.color_code
            ws['F5'] = order.pattern
            ws['A6'] = order_no
            ws['G6'] = ''  # Base
            ws['D6'] = order.order_qty  # order qty
            ws['C9'] = order.order_remark  # Remark

            order_no_split = order_no.split('-')
            
            qr_str = f"!BSVPD!{order_no_split[0]}!{order_no_split[1]}!"
            qr_img = qrcode.make(qr_str).resize((160, 160))
            
            # Save the QR code image to a BytesIO stream
            img_stream = io.BytesIO()
            qr_img.save(img_stream, format='PNG')
            
            # Reset the stream position to the beginning
            img_stream.seek(0)
            
            # Create an openpyxl image object
            qr_img_openpyxl = Image(img_stream)
            
            # Add the QR code image to the worksheet
            ws.add_image(qr_img_openpyxl, "G22")

            qr_img_2 = qrcode.make(qr_str).resize((180, 180))

            # Save the QR code image to a BytesIO stream
            img_stream_2 = io.BytesIO()
            qr_img_2.save(img_stream_2, format='PNG')
            
            # Reset the stream position to the beginning
            img_stream_2.seek(0)
            
            # Create an openpyxl image object
            qr_img_openpyxl_2 = Image(img_stream_2)
            
            # Add the QR code image to the worksheet
            ws.add_image(qr_img_openpyxl_2, "A22")


        # Save the workbook and return the response
        del wb['Sheet']
        wb.save(response)
        return response
    
    except Exception as e:
        # If an error occurs, return the error message in JSON format
        return JsonResponse({'error': str(e)})

# Upload swatch data to the database
@shared_task(bind=True, time_limit=1800)  # 30 minutes
def upload_swatch_data(self, file_path):
    """Upload swatch data to the database from .xlsb file"""
    progress_recorder = ProgressRecorder(self)
    error_logs = []
    success_count = 0
    update_count = 0
    error_count = 0
    
    try:
        # Read .xlsb file
        df = pd.read_excel(file_path, engine='pyxlsb', sheet_name='TOTAL SW')
        total_rows = len(df)
        
        # Required columns
        required_columns = ['EPC', 'CUSTOMER', 'M/S', 'STT', 'ITEM', 'COLOR', 'TYPE', 'BASE']
        
        # Check if all required columns exist
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            error_message = f"Missing required columns: {', '.join(missing_columns)}"
            logger.error(error_message)
            return {
                'status': 'error',
                'message': error_message
            }

        # Cập nhật progress bar trước khi bắt đầu xử lý
        progress_recorder.set_progress(0, total_rows, 'Starting data processing...')
        
        # Xử lý từng dòng
        for index, row in df.iterrows():
            try:
                # Check if row has all required values
                if row[required_columns].isnull().any():
                    error_logs.append(f"Row {index + 2}: Missing required values")
                    error_count += 1
                    continue

                # Clean and prepare data
                swatch_data = {
                    'epc': str(row['EPC']).strip(),
                    'stt': int(row['STT']),
                    'type': row['M/S'],
                    'customer': str(row['CUSTOMER']).strip(),
                    'item': str(row['ITEM']).strip(),
                    'color': str(row['COLOR']).strip(),
                    'pattern': str(row.get('TYPE', '')).strip(),
                    'base_color': str(row.get('BASE', '')).strip(),
                }

                # Process database in a separate transaction
                try:
                    with transaction.atomic():
                        # Check if swatch exists by EPC
                        existing_swatch = ColorSwatch.objects.filter(epc=swatch_data['epc']).first()
                        
                        if existing_swatch:
                            # Update existing swatch
                            for key, value in swatch_data.items():
                                setattr(existing_swatch, key, value)
                            existing_swatch.save()
                            update_count += 1
                        else:
                            # Create new swatch
                            ColorSwatch.objects.create(**swatch_data)
                            success_count += 1

                except Exception as db_error:
                    error_message = f"Row {index + 2}: Database error - {str(db_error)}"
                    error_logs.append(error_message)
                    logger.error(error_message)
                    error_count += 1

            except Exception as row_error:
                error_message = f"Row {index + 2}: {str(row_error)}"
                error_logs.append(error_message)
                logger.error(error_message)
                error_count += 1

            # Update progress bar after processing each row
            if (index + 1) % 10 == 0 or index == total_rows - 1:
                current_progress = index + 1
                progress_recorder.set_progress(
                    current_progress,
                    total_rows,
                    f'Processed {current_progress} of {total_rows} rows'
                )

        # Clean up temp file
        try:
            os.remove(file_path)
        except Exception as e:
            logger.warning(f"Failed to remove temp file: {str(e)}")

        # Final results
        result = {
            'status': 'success',
            'total_processed': total_rows,
            'new_records': success_count,
            'updated_records': update_count,
            'errors': error_count,
            'error_logs': error_logs[:100]  # Limit error logs to prevent response size issues
        }
        
        logger.info(f"Upload completed: {result}")
        return result

    except Exception as e:
        error_message = f"File processing error: {str(e)}"
        logger.error(error_message)
        return {
            'status': 'error',
            'message': error_message
        }

@app.task
def cleanup_old_swatch_movements():
    """Clean up old color swatch movement records"""
    deleted_count = ColorSwatchMovement.cleanup_old_records()
    logger.info(f"Cleanup task completed: {deleted_count} old records deleted")
    return deleted_count

# Register periodic task to run at midnight every day
app.conf.beat_schedule.update({
    'cleanup-old-swatch-movements': {
        'task': 'data_monitoring.tasks.cleanup_old_swatch_movements',
        'schedule': crontab(hour=0, minute=0),  # Run at midnight
    },
})